import os
import sys
import re
import datetime
import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


PDF_PATH = "ActiveReports Document.pdf"
HEADER_FONT_COLOR = "FFFFFF"
HEADER_FILL_COLOR = "4472C4"
COLUMN_WIDTH_NARROW = 20
COLUMN_WIDTH_WIDE = 25

SERVICE_PREFIXES = ['Ground', 'Next', '2nd', '3rd', '3 Day', 'Standard', 
                    'Worldwide', 'Express', 'Economy', 'Expedited']

SECTION_CONFIGS = {
    'drop_off': {
        'name': 'Drop Off Packages',
        'pattern': r'(?:Drop Off Packages\s+4x6 Pre-printed Shipping Label|Waybill \(Air/Ground Shipping Doc\)|Return Service \(UPS Only\))\s+Drop Off Date\s+Time Pickup Date Customer Service(.*?)(?=(?:Waybill \(Air/Ground|Return Service \(UPS Only\)|Enhanced Drop Off|Mobile Drop Off|Amazon\s+Drop Off|Page \d+ of \d+))',
        'tracking_patterns': [r'1Z[A-Z0-9]{16}']
    },
    'enhanced': {
        'name': 'Enhanced Drop Off Packages',
        'pattern': r'Enhanced Drop Off Packages.*?(?=Enhanced Drop Off Packages|Mobile Drop Off|Summary|Page \d+ of \d+)',
        'tracking_patterns': [r'1Z[A-Z0-9]{16}', r'HR[A-Z0-9]{6,8}-\d{4}-\d+']
    },
    'mobile': {
        'name': 'Mobile Drop Off',
        'pattern': r'Mobile Drop Off.*?(?=Summary|Kiosk Summary|$)',
        'tracking_patterns': [r'1Z[A-Z0-9]{16}']
    }
}


def extract_pdf_text(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            return '\n'.join(page.extract_text() or '' for page in pdf.pages)
    except FileNotFoundError:
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    except Exception as e:
        raise Exception(f"Error reading PDF: {e}")


def extract_report_info(pdf_path):
    """Extract report header and date range dynamically from first page of PDF"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            first_page_text = pdf.pages[0].extract_text()
            
            # Find any header line that contains a date range pattern
            lines = first_page_text.split('\n')
            for line in lines[:10]:  # Check first 10 lines
                # Look for date range pattern (DD MMM YYYY - DD MMM YYYY)
                date_match = re.search(r'(\d+ \w+ \d+ - \d+ \w+ \d+)', line)
                if date_match:
                    date_range = date_match.group(1)
                    # Extract everything before the date range as the header
                    header_text = line[:date_match.start()].strip()
                    if not header_text:
                        header_text = "Report"
                    return header_text, date_range
            
            return "Report", ""
    except Exception:
        return "Report", ""


def format_date_range(date_range_str):
    """Convert date range format from 'DD MMM YYYY - DD MMM YYYY' to 'DD-MM-YYYY to DD-MM-YYYY'"""
    parts = date_range_str.split(' - ')
    if len(parts) != 2:
        return date_range_str
    
    try:
        formatted_dates = []
        for date_str in parts:
            date_obj = datetime.datetime.strptime(date_str.strip(), '%d %b %Y')
            formatted_dates.append(date_obj.strftime('%d-%m-%Y'))
        return f"{formatted_dates[0]} to {formatted_dates[1]}"
    except ValueError:
        return date_range_str


def generate_output_filename(pdf_path, date_range):
    if date_range:
        pdf_name = pdf_path.replace('.pdf', '')
        formatted_date = format_date_range(date_range)
        return f"{pdf_name} {formatted_date}.xlsx".replace('  ', ' ')
    else:
        return pdf_path.replace('.pdf', '.xlsx')


def extract_service_types_from_pdf(pdf_path):
    service_types = set()
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:min(30, len(pdf.pages))]:
                text = page.extract_text()
                if not text:
                    continue
                
                for line in text.split('\n'):
                    if not (re.search(r'\d{1,2}:\d{2} (AM|PM)', line) and 
                           re.search(r'1Z[A-Z0-9]{16}', line)):
                        continue
                    
                    parts = line.split()
                    time_idx = next((i for i, p in enumerate(parts) if p in ['AM', 'PM']), None)
                    tracking = next((p for p in parts if p.startswith('1Z') and len(p) == 18), None)
                    
                    if not time_idx or not tracking:
                        continue
                    
                    date_start = time_idx + 1
                    date_count = sum(1 for i in range(date_start, min(date_start + 3, len(parts)))
                                   if i < len(parts) and not parts[i].startswith('1Z'))
                    
                    customer_start = time_idx + 1 + date_count
                    customer_end = parts.index(tracking)
                    
                    if customer_end > customer_start:
                        service_text = ' '.join(parts[customer_start:customer_end])
                        if any(service_text.startswith(p) for p in SERVICE_PREFIXES):
                            service_types.add(service_text)
        
        cleaned = {re.sub(r'\s*\d+\.?\d*lb$|\s*1Z[A-Z0-9]+$', '', s).strip() 
                   for s in service_types if 4 <= len(s) <= 50}
        return sorted(cleaned, key=len, reverse=True)
    
    except Exception as e:
        raise Exception(f"Error extracting service types: {e}")


def split_customer_service(customer_service, service_types):
    if not customer_service:
        return "", ""
    
    for service in service_types:
        if customer_service.endswith(service):
            customer = customer_service[:-len(service)].strip()
            return customer if customer else "", service
    
    return customer_service, ""


def extract_tracking(parts, tracking_patterns):
    for pattern in tracking_patterns:
        for part in parts:
            if re.match(pattern, part):
                return part
    return None


def extract_date_parts(parts, time_idx, num_parts=3):
    date_start = time_idx + 1
    date_parts = []
    for i in range(date_start, min(date_start + num_parts, len(parts))):
        if not any(parts[i].startswith(prefix) for prefix in ['1Z', 'HR']):
            date_parts.append(parts[i])
        else:
            break
    return date_parts


def extract_weight(parts):
    if parts[-1] in ['ManWt', 'Com']:
        if len(parts) >= 2 and (any(x in parts[-2] for x in ['lb', 'kg']) or parts[-2] == 'N/A'):
            return f"{parts[-2]} {parts[-1]}"
        return 'N/A'
    return parts[-1]


def parse_line_entry(line, service_types, tracking_patterns):
    parts = line.split()
    if len(parts) < 4:
        return None
    
    time_idx = next((i for i, p in enumerate(parts) if p in ['AM', 'PM']), None)
    if time_idx is None:
        return None
    
    tracking = extract_tracking(parts, tracking_patterns)
    if not tracking:
        return None
    
    time = ' '.join(parts[0:time_idx + 1])
    date_parts = extract_date_parts(parts, time_idx)
    pickup_date = ' '.join(date_parts) if date_parts else ''
    weight = extract_weight(parts)
    
    customer_start = time_idx + 1 + len(date_parts)
    customer_end = parts.index(tracking)
    customer = ' '.join(parts[customer_start:customer_end]) if customer_end > customer_start else ''
    
    customer_name, service = split_customer_service(customer, service_types)
    
    return {
        'Time': time,
        'Pickup Date': pickup_date,
        'Customer': customer_name,
        'Service': service,
        'Tracking Number': tracking,
        'Weight': weight
    }


def parse_section(text, section_config, service_types):
    sections = re.findall(section_config['pattern'], text, re.DOTALL)
    data = []
    
    for section in sections:
        for line in section.split('\n'):
            entry = parse_line_entry(line, service_types, section_config['tracking_patterns'])
            if entry:
                data.append(entry)
    
    return data


def parse_tabular_section(text, section_name):
    if section_name == "Summary":
        pattern = r'Summary\s+Unknown service.*?Type\s+International\s+Ground\s+Air\s+.*?Total Packages\s+Weight\s+(.*?)(?=Kiosk Summary|Page \d+ of \d+)'
    else:
        pattern = r'Kiosk Summary.*?Type\s+International\s+Ground\s+Air\s+.*?Total Packages\s+Weight\s+(.*?)(?=Page \d+ of \d+)'
    
    match = re.search(pattern, text, re.DOTALL)
    if not match:
        return []
    
    data = []
    lines = match.group(1).strip().split('\n')
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        
        if not line or line.startswith('Total:'):
            i += 1
            continue
        
        if re.match(r'^\d+\.\d+\.\d+\.\d+$', line) or re.match(r'^\(\w{2}-\w{2}\)', line):
            i += 1
            continue
        
        if i + 1 < len(lines) and re.match(r'^[\d.,]+$', lines[i + 1].strip()):
            next_val = lines[i + 1].strip()
            if '.' in next_val or float(next_val.replace(',', '')) < 100000:
                line = f"{line} {next_val}lb"
                i += 2
            else:
                i += 1
        else:
            i += 1
        
        match = re.search(r'^(.*?)\s+(\d+)\s+([\d,]+)\s+(\d+)\s+(\d+)\s+([\d,]+)\s*([\d.,]+(?:lb|kg)?)?', line)
        
        if match:
            type_desc = match.group(1).strip()
            if type_desc and type_desc != 'Total:':
                data.append({
                    'Type': type_desc,
                    'International': match.group(2),
                    'Ground': match.group(3).replace(',', ''),
                    'Air': match.group(4),
                    'Unknown service, not specified': match.group(5),
                    'Total Packages': match.group(6).replace(',', ''),
                    'Weight': match.group(7) if match.group(7) else ''
                })
    
    return data


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR)
        cell.fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_worksheet(workbook, sheet_name, data, column_width=COLUMN_WIDTH_WIDE):
    if not data:
        return
    
    ws = workbook.create_sheet(sheet_name)
    df = pd.DataFrame(data).fillna('')
    
    if sheet_name in ['Summary', 'Kiosk Summary']:
        numeric_columns = ['International', 'Ground', 'Air', 'Unknown service, not specified', 'Total Packages']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    style_header(ws)
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = column_width


def create_excel_file(output_filename, sheets_data):
    wb = Workbook()
    wb.remove(wb.active)
    
    for sheet_name, data, width in sheets_data:
        create_worksheet(wb, sheet_name, data, width)
    
    wb.save(output_filename)


def main(pdf_path=PDF_PATH):
    try:
        print(f"📄 Reading PDF: {pdf_path}")
        service_types = extract_service_types_from_pdf(pdf_path)
        print(f"✓ Found {len(service_types)} service types")
        
        report_header, date_range = extract_report_info(pdf_path)
        output_filename = generate_output_filename(pdf_path, date_range)
        if date_range:
            print(f"✓ Report: {report_header} {date_range}")
        else:
            print(f"✓ Report: {report_header}")
        
        print("\n📖 Extracting text from PDF...")
        full_text = extract_pdf_text(pdf_path)
        
        print("\n⚙️  Processing Drop Off Packages...")
        drop_off_data = parse_section(full_text, SECTION_CONFIGS['drop_off'], service_types)
        print(f"✓ Extracted {len(drop_off_data)} entries")
        
        print("\n⚙️  Processing Enhanced Drop Off Packages...")
        enhanced_data = parse_section(full_text, SECTION_CONFIGS['enhanced'], service_types)
        print(f"✓ Extracted {len(enhanced_data)} entries")
        
        print("\n⚙️  Processing Mobile Drop Off...")
        mobile_data = parse_section(full_text, SECTION_CONFIGS['mobile'], service_types)
        print(f"✓ Extracted {len(mobile_data)} entries")
        
        print("\n⚙️  Processing Summary sections...")
        summary_data = parse_tabular_section(full_text, "Summary")
        kiosk_data = parse_tabular_section(full_text, "Kiosk Summary")
        print(f"✓ Extracted {len(summary_data)} Summary entries")
        print(f"✓ Extracted {len(kiosk_data)} Kiosk Summary entries")
        
        sheets_data = [
            ("Drop Off Packages", drop_off_data, COLUMN_WIDTH_NARROW),
            ("Enhanced Drop Off Packages", enhanced_data, COLUMN_WIDTH_WIDE),
            ("Mobile Drop Off", mobile_data, COLUMN_WIDTH_WIDE),
            ("Summary", summary_data, COLUMN_WIDTH_WIDE),
            ("Kiosk Summary", kiosk_data, COLUMN_WIDTH_WIDE),
        ]
        
        print("\n📊 Creating Excel file...")
        create_excel_file(output_filename, sheets_data)
        print(f"\n✅ Success! Excel file created: {output_filename}")
        print(f"   Total packages extracted: {len(drop_off_data) + len(enhanced_data) + len(mobile_data)}")
        
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    pdf_path = sys.argv[1] if len(sys.argv) > 1 else PDF_PATH
    main(pdf_path)